#!/usr/bin/env python3
"""Generate the Yale (YNHHS) Data Remote-Hosted Security Design Review responses
as an .xlsx workbook reflecting the ImPACTS PECC Support Tool's implemented controls.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DESKTOP = os.path.expanduser("~/Desktop")
OUT = os.path.join(DESKTOP, "Yale Security Design Review - ImPACTS Responses.xlsx")

# Each row: (sec, question, answer)  answer in {"Yes","No","N/A",""}; comment
# Tuple form: (sec, question, answer, comment)
SECTION = "SECTION"  # marker for section header rows

rows = [
    ("Vendor Name", "- Dr. Marc Auerbach, MD, FAP, MSc. Professors of Pediatrics (Emergency Medicine) and of Emergency Medicine, and\n- Benjamin Michaels, MBA, MPH (B.L. Michaels, LLC) — contractor working with the Yale School of Medicine, Pediatrics Department", "", ""),
    ("Application and Version", "PECC Support Tool\n(web application; continuously deployed / rolling releases)", "", ""),
    ("Date Questionnaire Completed", "06/24/2026", "", ""),

    (SECTION, "SEC 0 — General & Organizational Questions", "", ""),
    ("SEC 0.a", "Do you have Cyber security insurance?", "No",
     "No active policy."),
    ("SEC 0.b", "If yes, what are your Policy limits?", "N/A",
     "No active policy."),
    ("SEC 0.c", "Is your Policy Per-claim or Per-occurrence?", "N/A",
     "No active policy."),
    ("SEC 0.d", "Does your system/application support multi-factor authentication for administrative access to systems and its supporting infrastructure?", "Yes",
     "Yes. Admin (platform admin) accounts authenticate via Supabase Auth with mandatory TOTP MFA (authenticator app) plus 15+ character password. Same enforced MFA gate as all users."),
    ("SEC 0.e", "Does your system/application support multi-factor authentication for general user access?", "Yes",
     "Yes — mandatory for all users. TOTP MFA (Google Authenticator, 1Password, Authy, Apple Passwords, etc.) must be enrolled at first sign-in via a full-screen security gate before any app content loads; users are re-challenged at login and when resuming a session after inactivity."),
    ("SEC 0.f", "Will this system/application be using any unsupported or End of Life software/hardware? If yes, timeline to replace/upgrade?", "No",
     "No. Current stack: React 18, Material UI, supabase-js, PostgreSQL 17, Node.js LTS. Dependencies monitored (Dependabot)."),
    ("SEC 0.g", "Will this system/application be using any Open Source software? (list software)", "Yes",
     "Yes. React, React Router, Material UI (MUI), @supabase/supabase-js, Deno (edge functions). Backend Postgres (PostgreSQL 17). Hosted on managed platforms (Supabase, Vercel)."),
    ("SEC 0.h", "Are there any web site URLs used as part of this implementation? (include all URLs)", "Yes",
     "Primary: https://peccsupporttool.com . Legacy https://impacts-tau.vercel.app permanently redirects (308) to primary. Backend API: https://ftpifgzzfwpujlvbqqhu.supabase.co (Supabase)."),
    ("SEC 0.i", "Are any of the web sites hosted external to the United States (Overseas)?", "No",
     "No. All hosting is US-based (see 0.j)."),
    ("SEC 0.j", "Is data hosted only within the U.S.? (if so where?)", "Yes",
     "Yes. Database/Auth on Supabase running on AWS region us-east-2 (Ohio, USA). Web app served via Vercel (US). No overseas storage."),
    ("SEC 0.k", "If your device requires a wireless connection, which wireless frequencies/standards does it support (802.11 a/b/g/n)?", "N/A",
     "N/A. Browser-based SaaS web application; no vendor-supplied hardware or wireless device."),

    (SECTION, "SEC 1 — System / Application and Data Overview", "", ""),
    ("SEC 1.a", "Does the vendor utilize a public, 3rd party data-hosting service provider (AWS, Azure, RackSpace, etc.)?", "Yes",
     "Yes. Supabase (managed PostgreSQL, Auth, Edge Functions) running on Amazon Web Services (AWS). Vercel for web application hosting/CDN."),
    ("SEC 1.b", "Have the vendor remote-hosted environment / supporting systems attained a 3rd party assessment (e.g. SSAE-16 SOC Type audit)? Specify type & date.", "No",
     "ImPACTS Collaborative has NOT independently undergone a SOC audit. Underlying infrastructure providers maintain SOC 2 / ISO 27001 (AWS; Vercel; Supabase on Team/Enterprise plans). ImPACTS is currently on Supabase Free tier."),
    ("SEC 1.c", "Is periodic penetration testing conducted on environment and underlying systems? If yes, how often?", "No",
     "No formal scheduled third-party penetration testing by ImPACTS at this time. Application-level controls in place: RLS, mandatory MFA, CSP and security headers, input sanitization. Underlying providers perform their own platform testing."),
    ("SEC 1.d", "Do you have an enterprise-wide EDR solution implemented in the corporate environment? Confirm which EDR and coverage.", "No",
     "CONFIRM (business): No enterprise EDR fleet managed by ImPACTS Collaborative to our knowledge."),
    ("SEC 1.e", "Is there an EDR solution implemented in the vendor-hosted platform storing/handling YNHH data? Confirm EDR & coverage.", "N/A",
     "N/A. Fully managed PaaS (Supabase/Vercel on AWS); no vendor-managed servers/OS on which to install EDR. Host/OS security is managed by the cloud providers."),

    (SECTION, "SEC 2 — Data Privacy", "", ""),
    ("SEC 2.a", "Do application/supporting systems handle PHI, PII, PCI, Financial, Regulatory or other critical data? Specify type.", "Yes",
     "Limited business-contact PII only. NO PHI, NO PCI, NO financial data. Collects professional/account data for clinical staff users. PHI is explicitly PROHIBITED by the Terms of Service ('zero tolerance' for patient data)."),
    ("SEC 2.b", "If app handles PHI/PII/PCI/critical data, list identifiers disclosed to or accessed by your workforce.", "",
     "Identifiers collected: user first/last name, work email, phone number, job title, department, hospital/health-system affiliation. NO patient identifiers (no MRN, SSN, DOB, addresses, clinical data)."),
    ("SEC 2.c", "If the application handles PHI, is there a Business Associate Agreement (BAA) in place? If no, explain.", "N/A",
     "N/A by design — the Tool prohibits PHI, so no BAA is required for intended use. If YNHHS requires a BAA as a condition, ImPACTS can pursue one separately (Supabase/Vercel offer BAAs on paid/enterprise plans)."),
    ("SEC 2.d", "Are we sending very sensitive health information (HIV/AIDS, psychiatric, genetic, substance use)? Specify.", "No",
     "No. No clinical or patient health information of any kind is collected or transmitted."),
    ("SEC 2.e", "Will your organization have accounts (e.g., support accounts) that can access the YNHH instance/data?", "Yes",
     "Yes. ImPACTS platform administrators can access application data for operations/support."),
    ("SEC 2.f", "If yes, are these accounts shared or unique accounts assigned to individuals?", "Unique",
     "Unique. Every administrator has an individual, named account. No shared/generic accounts."),
    ("SEC 2.g", "If assigned to individuals, what authentication mechanisms restrict access (AD login, MFA, etc.)?", "",
     "Individual Supabase Auth credentials: unique email + 15+ character password + mandatory TOTP MFA. Role enforced server-side via PostgreSQL Row-Level Security (RLS)."),
    ("SEC 2.h", "Does your organization have an internal ticketing system to manage onboarding/offboarding?", "No",
     "CONFIRM (business): No formal ticketing system. In-app admin controls allow immediate account provisioning/deactivation."),
    ("SEC 2.i", "If yes, how soon is access removed after an employee is offboarded?", "",
     "Access can be revoked immediately by an admin (account deactivation in the Team admin panel) and/or by disabling the Supabase user. Vendor to confirm formal SLA."),
    ("SEC 2.j", "Does your organization perform access reviews of critical systems on a consistent schedule?", "Yes",
     "Tooling exists: admins can export a User Access / Entitlement report (CSV) for periodic reviews. Formal review cadence to be finalized (recommended quarterly)."),
    ("SEC 2.k", "If yes, how often are access reviews performed?", "",
     "Recommended quarterly. CSV entitlement export (name, email, role, platform-admin flag, status, last login, created date) supports the review."),
    ("SEC 2.l", "How many users from your company will have access to YNHHS data?", "",
     "1-10 (ImPACTS administrators)."),
    ("SEC 2.m", "Will you use subcontractors with access to YNHHS data? Do you have a BAA with them?", "Yes",
     "Subprocessors (infrastructure only): Supabase (DB/Auth/functions), Vercel (web hosting), Resend (transactional email). They process account/operational data, not PHI. BAAs available via Supabase/Vercel paid plans if required."),
    ("SEC 2.n", "How long is data retained?", "",
     "Account and operational data retained for the life of the account/program. Audit and security-event logs retained per policy (recommended 6 years for audit trail). Configurable."),
    ("SEC 2.o", "How is YNHHS data purged/destroyed?", "",
     "On account/data deletion request, records are removed from the PostgreSQL database (logical delete followed by hard delete). Backups age out per provider backup retention."),
    ("SEC 2.p", "Will your company provide attestation of appropriate disposal of data at the time of disposition?", "Yes",
     "Yes — attestation can be provided upon request."),
    ("SEC 2.q", "Does your organization have policies/procedures related to handling PHI/PII/PCI/critical data?", "Yes",
     "Documented data-use restrictions in the Terms of Service prohibit PHI/PII patient data and define acceptable use, logging, and access. Formal standalone policy document can be provided/expanded on request."),
    ("SEC 2.r", "Are staff provided training on the policies/procedures?", "No",
     "CONFIRM (business): No formal recurring training program documented. Vendor to confirm."),

    (SECTION, "SEC 3 — Application User-Access and Controls", "", ""),
    ("SEC 3.a", "If installed with default accounts, can they be modified/removed (e.g. 'admin','guest')?", "Yes",
     "Yes. No generic default/built-in accounts (e.g. admin, guest). Public self-registration is disabled; all accounts are created only through a validated invitation link issued by an authorized administrator."),
    ("SEC 3.b", "Does the system/application prohibit the provisioning/use of shared accounts?", "Yes",
     "Yes. Each account requires a unique email and individual credentials. Access is invitation-only with role-based provisioning; shared or generic logins are not permitted."),
    ("SEC 3.c", "Can user access management be conducted by a YNHHS Application Owner/System Administrator? If no, who manages?", "No",
     "Currently managed by ImPACTS administrators (no per-customer admin delegation yet). A designated YNHHS admin role could be provisioned on request."),
    ("SEC 3.d", "Does the system/application interface with enterprise directory services (AD, LDAP, ADFS) for authentication?", "No",
     "Not at this time. SSO/SAML (e.g., Azure AD/Entra) is on the roadmap (Supabase supports SAML on paid tiers) but not yet implemented. Authentication is native email/password + MFA. (NOTE: if SSO is later integrated, 3.f-3.n become N/A under AD policy.)"),
    ("SEC 3.e", "Does account provisioning support RBAC (Role Based Access Control)?", "Yes",
     "Yes. Roles: admin, manager, mentor, PECC, hospital_system, hiring_group. Enforced server-side via PostgreSQL Row-Level Security (RLS) policies, not just UI."),
    ("SEC 3.f", "Do all user accounts utilize at least username/password for authentication?", "Yes",
     "Yes. Email (username) + password, plus mandatory TOTP MFA."),
    ("SEC 3.g", "Can access controls employ Multi-Factor Authentication? Specify criteria.", "Yes",
     "Yes — mandatory TOTP MFA (RFC 6238 time-based one-time passwords) via any standard authenticator app. Enforced via full-screen gate before app access; re-challenged on login and after inactivity."),
    ("SEC 3.h", "Does the system/application prohibit user accounts without passwords (no blank passwords)?", "Yes",
     "Yes. Minimum 15-character password enforced at four layers (Supabase Auth config, client, edge functions, legacy-user upgrade prompt). Blank/whitespace-only rejected."),
    ("SEC 3.i", "Does the system/application require unique user accounts?", "Yes",
     "Yes. Unique email per account."),
    ("SEC 3.j", "Does the system/application enforce password complexity? Specify criteria.", "Yes",
     "Yes. Minimum 15 characters (NIST 800-63B length-first guidance), with guidance/checks for uppercase, lowercase, number, and symbol; mostly-whitespace passwords rejected. Supabase enforces minimum length server-side. Note: compromised-password screening (HaveIBeenPwned) is available but not currently enabled (paid-tier feature); will enable on upgrade to Supabase Pro."),
    ("SEC 3.k", "Does the system/application allow password complexity to be customized?", "Yes",
     "Yes. Policy is configurable in the Supabase auth config and the application password-policy module."),
    ("SEC 3.l", "Does the system/application enforce a password expiration interval? Specify.", "No",
     "No forced periodic expiration (aligned with NIST 800-63B, which discourages routine rotation). However, users whose passwords predate the current 15-char policy are force-prompted to upgrade at login."),
    ("SEC 3.m", "Does the system/application enforce password history requirements? Specify.", "No",
     "No password-reuse history enforcement currently (not enforced by Supabase Auth by default)."),
    ("SEC 3.n", "Does the system/application enforce minimum password age? Specify.", "No",
     "No minimum password age enforced."),
    ("SEC 3.o", "Are account password resets self-managed (without admin assistance)?", "Yes",
     "Yes. Self-service 'Forgot password?' email recovery flow. (Reset link flow itself does not require MFA; MFA applies again at normal sign-in.)"),
    ("SEC 3.p", "Do account password resets require assistance from a system administrator? Explain.", "No",
     "No. Resets are self-service via email. Admins can also trigger resets if needed."),
    ("SEC 3.q", "Are account lock-out controls for failed password attempts in place? Specify.", "Yes",
     "Supabase Auth applies rate limiting/throttling on authentication endpoints to deter brute force. Every failed login is recorded in the append-only security_events log (email, timestamp, reason). No fixed hard-lockout threshold is exposed to clients."),
    ("SEC 3.r", "Do session time-out controls force re-authentication due to inactivity? Specify.", "Yes",
     "Yes. 30-minute idle auto sign-out (shared-workstation safeguard), synchronized across browser tabs; idle timeouts are logged. JWT access tokens expire hourly with refresh-token rotation."),
    ("SEC 3.s", "Do session time-outs black out or hide state of activity on screens/monitors?", "Yes",
     "Yes. On idle timeout the user is signed out and returned to the login screen. A full-screen security gate hides all application content whenever the session is not fully authenticated (password/terms/MFA pending)."),

    (SECTION, "SEC 4 — Vendor Remote-Hosted Environment / Application Controls", "", ""),
    ("SEC 4.a", "Is client data logically segregated from all other client data in the vendor remote-hosted environment?", "Yes",
     "Yes — logically. Multi-tenant PostgreSQL database with Row-Level Security (RLS) enforced on all 50 tables (verified), scoping every query by per-row ownership and role. (Logical segregation via RLS, not a separate DB instance per client.) Note: one public storage bucket (program-logos) holds non-sensitive program logo images with public read; no client/user data is stored there."),
    ("SEC 4.b", "Does the environment have network controls restricting direct public access to system components?", "Yes",
     "Yes. The database is not directly publicly exposed; access is via authenticated API (PostgREST/Supabase) gated by API keys, signed JWTs, and RLS. Web tier served via Vercel edge. Managed by Supabase/AWS/Vercel."),
    ("SEC 4.c", "Does the environment have host-based controls (host firewalls, IPS/IDS) to prevent unauthorized access?", "Yes",
     "Yes — provided and managed by the cloud platforms (AWS/Supabase/Vercel). ImPACTS does not manage host OS."),
    ("SEC 4.d", "Do vendor remote-hosted systems have anti-virus mechanisms current and actively running?", "N/A",
     "N/A. Serverless/managed PaaS; no vendor-managed servers/OS. Host-level protection is the responsibility of AWS/Supabase/Vercel."),
    ("SEC 4.e", "Does the environment routinely implement patching/updates in a timely manner?", "Yes",
     "Yes. Underlying platforms (Supabase/Vercel/AWS) are continuously patched by the providers. Application dependencies are monitored and updated (Dependabot)."),
    ("SEC 4.f", "Does the vendor require remote access (e.g., VPN) for support? Define requirements.", "No",
     "No VPN/remote-into-customer-environment. SaaS model. ImPACTS admins operate via the authenticated web app and the Supabase dashboard (MFA-protected)."),
    ("SEC 4.g", "Does remote access support unattended or attended (YNHH manual intervention) access?", "N/A",
     "N/A. No remote access into any YNHHS environment; multi-tenant SaaS."),
    ("SEC 4.h", "Are data files stored on vendor servers/databases stored encrypted? Specify (AES-128/256).", "Yes",
     "Yes. Data at rest is encrypted (AES-256) by Supabase/AWS managed storage."),
    ("SEC 4.i", "If data files must be stored on vendor mobile/handheld devices, is data encrypted? Specify.", "N/A",
     "N/A. No native mobile app and no local device data store."),
    ("SEC 4.j", "If wireless communications occur with devices/modalities, do transmissions use strong encryption? Specify.", "N/A",
     "N/A. No device/modality wireless communication."),
    ("SEC 4.k", "If web-browser sessions interact with the app, do transmissions use strong encryption? Specify.", "Yes",
     "Yes. All traffic over HTTPS/TLS 1.2+ (TLS 1.3 supported). HSTS via 'upgrade-insecure-requests'; strict Content-Security-Policy; X-Frame-Options: DENY; X-Content-Type-Options: nosniff."),
    ("SEC 4.l", "Is a mobile application component in-scope for YNHH use?", "No",
     "No. Responsive web application only (works in mobile browsers; no installed mobile app)."),
    ("SEC 4.m", "If yes, is data retained locally on the user's mobile device?", "N/A",
     "N/A. No mobile app."),
    ("SEC 4.n", "If yes, how are data transmissions between the mobile app and web app established?", "N/A",
     "N/A. No mobile app."),
    ("SEC 4.o", "If yes, how is this data transmission secured?", "N/A",
     "N/A. No mobile app."),
    ("SEC 4.p", "Does your system/application require SMBv1?", "No", "No."),
    ("SEC 4.q", "Does your system/application require Java? If yes, which version?", "No",
     "No. JavaScript/TypeScript (React) front end; Deno/TypeScript edge functions. No Java runtime."),
    ("SEC 4.r", "Does your system/application require a mail relay from clients?", "No",
     "No. Outbound transactional email (invitations, password resets) is sent via Resend; no client mail relay required."),
    ("SEC 4.s", "Is your system/application IPv6 capable?", "No",
     "Verified 06/24/2026 via public DNS (1.1.1.1, 8.8.8.8): neither peccsupporttool.com nor the Supabase API hostname (ftpifgzzfwpujlvbqqhu.supabase.co) publishes AAAA (IPv6) records — both resolve to IPv4 only (Vercel/Cloudflare edge). The application is reached by hostname over HTTPS; clients on IPv6-only networks still connect via standard dual-stack / NAT64/DNS64. Native IPv6 endpoints are not currently published."),
    ("SEC 4.t", "What operating system is required for servers, clients, etc.? Specify.", "",
     "No client OS requirement — a modern web browser (Chrome, Edge, Safari, Firefox) only. Server side is fully managed by Supabase/Vercel (Linux-based); not customer-managed."),
    ("SEC 4.u", "Is your system/application compatible with MS Defender for Endpoint virus scanning? Specify.", "Yes",
     "Yes/Compatible. Browser-based app with no installed agent; compatible with any endpoint protection including Microsoft Defender for Endpoint."),
    ("SEC 4.v", "Do any URLs need to be whitelisted? Specify.", "Yes",
     "Allow: https://peccsupporttool.com , https://*.supabase.co (specifically https://ftpifgzzfwpujlvbqqhu.supabase.co and wss://ftpifgzzfwpujlvbqqhu.supabase.co)."),
    ("SEC 4.w", "Do any IP addresses need to be whitelisted? Specify source/dest IP:Port:Protocol.", "No",
     "No fixed IPs to whitelist — Vercel/Supabase use dynamic edge IPs. Use hostnames over HTTPS (443/TCP)."),

    (SECTION, "SEC 5 — API Controls", "", ""),
    ("SEC 5.a", "Do call transmissions utilize strong encryption protocols? Specify type.", "Yes",
     "Yes. All API calls over HTTPS/TLS 1.2+ (TLS 1.3 supported)."),
    ("SEC 5.b", "Do transmissions undergo end-to-end authentication (request to termination)? Specify method (SAML/OAUTH).", "Yes",
     "Yes. OAuth2-style bearer authentication using Supabase-issued signed JWTs. Short-lived access tokens (1 hour) with refresh-token rotation."),
    ("SEC 5.c", "Are controls in place to sign, hash, timestamp, and strictly scope tokens (prevent replay/tampering/escalation)?", "Yes",
     "Yes. JWTs are cryptographically signed with 1-hour expiry; refresh-token rotation with a reuse-interval guard detects token reuse. PostgreSQL RLS scopes every request to the caller's user/role, limiting accessible data per token."),
    ("SEC 5.d", "Are controls in place to limit API access requests (prevent brute-forcing)?", "Yes",
     "Yes. Supabase platform rate-limits auth endpoints; sensitive edge functions (e.g., invitation registration) enforce DB-backed rate limiting."),
    ("SEC 5.e", "Are controls in place to prevent manipulation of input/output handling (SQLi, XSS)?", "Yes",
     "Yes. Parameterized queries via PostgREST/supabase-js (no string-built SQL); RLS as defense-in-depth; React output auto-escaping; strict CSP, X-Content-Type-Options: nosniff, X-Frame-Options: DENY, X-XSS-Protection."),

    (SECTION, "SEC 6 — Monitoring, Tracking, and Auditing", "", ""),
    ("SEC 6.a", "Can the system provide a report log of active/inactive YNHHS user/service accounts for entitlement reviews?", "Yes",
     "Yes. Admin 'User Access / Entitlement' CSV export: first/last name, email, phone, role, platform-admin flag, status (active/inactive), reports-to, last login, account-created date."),
    ("SEC 6.b", "Can the system provide a report log of active/inactive vendor user/service accounts accessing YNHHS data?", "Yes",
     "Yes. ImPACTS admin accounts appear in the same user list/entitlement export."),
    ("SEC 6.c", "Does the system retain a log of access account creations, modifications and deletions?", "Yes",
     "Yes. Append-only audit_log (HIPAA-oriented) records create/update/delete on sensitive tables (users, hospitals, contacts, CRM) with old/new values."),
    ("SEC 6.d", "Does the account creation/modification log include time/date?", "Yes",
     "Yes. Each audit record has a created_at timestamp (UTC)."),
    ("SEC 6.e", "Does the log include the UserID/Service Account that executed the action?", "Yes",
     "Yes. The acting user is captured (performed_by / user id)."),
    ("SEC 6.f", "Does the system retain a log of a UserID's last successful login session?", "Yes",
     "Yes. Successful sign-in events are recorded (last-login tracked per user; SIGNED_IN events)."),
    ("SEC 6.g", "Does the system retain a log of a vendor UserID's last successful login used to access YNHHS data?", "Yes",
     "Yes. Same login logging applies to ImPACTS admin accounts."),
    ("SEC 6.h", "Does the system retain a log of a UserID's last failed login attempt?", "Yes",
     "Yes. Failed logins recorded in append-only security_events (email, timestamp, reason)."),
    ("SEC 6.i", "Does the system retain a log of a vendor UserID's last failed login attempt?", "Yes",
     "Yes. Same failed-login logging applies to ImPACTS admin accounts."),
    ("SEC 6.j", "Does the system log activity of accounts that view/modify/remove PHI or sensitive data?", "N/A",
     "N/A for PHI — no PHI is stored. Sensitive operational records (users, hospitals, contacts, CRM) have create/update/delete audited via audit_log; security events (login, MFA, idle timeout, password changes) logged separately."),
    ("SEC 6.k", "Does the system log activity of vendor accounts that view/modify/remove PHI or sensitive data?", "N/A",
     "N/A for PHI — none stored. Vendor (admin) modifications to sensitive operational records are captured in audit_log with actor and timestamp."),
    ("SEC 6.l", "Are all logs configured so they cannot be changed or deleted?", "Yes",
     "Yes. audit_log and security_events are append-only — UPDATE and DELETE are revoked from application (anon/authenticated) roles. INSERT into security_events is intentionally permitted to clients so that pre-authentication events (e.g., failed logins) can be recorded; records still cannot be modified or removed. SELECT on these logs is admin-only."),
    ("SEC 6.m", "Are logs stored on the recording server or pushed to a centralized location (SIEM)?", "No",
     "Stored in dedicated append-only tables within the managed Supabase PostgreSQL database (separate from operational logic). No external SIEM integration currently; Supabase platform logs also available. SIEM/log-drain export available on Supabase paid tiers."),
    ("SEC 6.n", "Are all logs maintained with unlimited retention and available on request in reportable format?", "Yes",
     "Application audit_log/security_events persist in the database and are reportable via SQL/CSV (retention configurable; recommended 6 years). NOTE: Supabase *platform* log retention is tier-limited (1 day on Free; 7+ days on paid). Application audit tables themselves are retained indefinitely until purged per policy."),

    (SECTION, "SEC 7 — Disaster Recovery and Business Continuity", "", ""),
    ("SEC 7.a", "Does the environment have DR/BC measures for continuous service/data availability? Specify.", "Yes",
     "Hosted on AWS (Supabase) and Vercel, which provide infrastructure redundancy and high availability. IMPORTANT: ImPACTS is currently on the Supabase Free tier, which does NOT include automated daily backups. Upgrading to Supabase Pro enables daily backups (7-day retention); Point-in-Time Recovery is available as an add-on. Recommend upgrading to Pro before production YNHHS use."),
    ("SEC 7.b", "Does the environment have DR limitations that could affect client RTO/RPO objectives? Explain.", "Yes",
     "Yes — on the current Free tier there are no managed/automated backups, so RPO/RTO cannot be guaranteed. This is remediated by upgrading to Supabase Pro/Team plus the Point-in-Time Recovery add-on, which we will do for production YNHHS deployment."),

    (SECTION, "SEC 8 — Vendor Education and Training", "", ""),
    ("SEC 8.a", "Does your organization provide information security training to its employees?", "No",
     "CONFIRM (business): No formal recurring InfoSec training program documented. Vendor to confirm/establish."),
    ("SEC 8.b", "If yes, is social engineering content included in the training?", "N/A",
     "N/A pending a formal training program."),
    ("SEC 8.c", "If yes, is password management and security content included?", "N/A",
     "N/A pending a formal training program. (Product enforces strong passwords + MFA technically.)"),
    ("SEC 8.d", "If yes, are organizational data classifications included?", "N/A", "N/A pending a formal training program."),
    ("SEC 8.e", "If yes, are asset management expectations included?", "N/A", "N/A pending a formal training program."),
    ("SEC 8.f", "If yes, how often is training provided?", "N/A", "N/A pending a formal training program."),
    ("SEC 8.g", "If yes, how is the training delivered to employees?", "N/A", "N/A pending a formal training program."),
    ("SEC 8.h", "Are phishing simulations created by your InfoSec team and delivered to employees?", "No",
     "CONFIRM (business): No phishing simulation program currently."),
    ("SEC 8.i", "If yes, how often are the 'fake' phishing emails sent?", "N/A", "N/A."),
    ("SEC 8.j", "If yes, what is the current click percentage for your organization?", "N/A", "N/A."),

    (SECTION, "SEC 9 — Incident Response", "", ""),
    ("SEC 9.a", "Does your organization have a formal enterprise Incident Response Plan?", "No",
     "CONFIRM (business): No formal documented enterprise IR plan at this time. Vendor to confirm/establish. Application provides audit + security-event logging to support investigation."),
    ("SEC 9.b", "If yes, how often is the plan reviewed and/or updated?", "N/A", "N/A pending a formal IR plan."),
    ("SEC 9.c", "Does your organization have a dedicated Incident Response contact responsible for the response process?", "Yes",
     "Primary contact: ImPACTS Collaborative — impactscollaborative@gmail.com."),
    ("SEC 9.d", "If yes, who is this individual (name and email)?", "",
     "ImPACTS Collaborative administrator — impactscollaborative@gmail.com. (Provide named individual on the form.)"),
    ("SEC 9.e", "Is there a 24/7/365 staffed phone line/email dedicated to incident response?", "No",
     "No 24/7 staffed line. Email contact monitored: impactscollaborative@gmail.com."),
    ("SEC 9.f", "If yes, what is the phone/email contact information?", "",
     "impactscollaborative@gmail.com (email; not 24/7)."),
    ("SEC 9.g", "Has your organization developed playbooks for common incident types (phishing, ransomware, etc.)?", "No",
     "CONFIRM (business): No formal playbooks currently."),

    (SECTION, "SEC 10 — Artificial Intelligence (complete only if product uses AI)", "", ""),
    ("SEC 10.a", "How will the data be secured, transmitted and stored?", "N/A",
     "N/A. The ImPACTS PECC Support Tool does not use AI/ML in the application; no end-user data is processed by an AI model."),
    ("SEC 10.b", "How do you plan on mitigating risk and bias?", "N/A", "N/A. No AI component in the product."),
    ("SEC 10.c", "What content is being generated (code, text, images, other)?", "N/A", "N/A. No AI component in the product."),
    ("SEC 10.d", "What Yale New Haven Systems data does the AI system/tool require access to?", "N/A", "N/A. No AI component; no YNHHS data used by any AI."),
    ("SEC 10.e", "What model is being used?", "N/A", "N/A. No AI model in the product."),
    ("SEC 10.f", "Who can modify/update the AI model?", "N/A", "N/A. No AI model in the product."),
    ("SEC 10.g", "What security mechanisms identify and filter malicious inputs?", "N/A",
     "N/A for AI. General input hardening (parameterized queries, RLS, CSP, output escaping) applies to the application."),
    ("SEC 10.h", "Will the AI model be trained using our data?", "No",
     "No. No AI model; YNHHS data will not be used to train any model."),

    (SECTION, "SEC 11 — Additional Information", "", ""),
    ("SEC 11.a", "Are there additional security controls that you would like to highlight?", "Yes",
     "Highlights: (1) MANDATORY TOTP MFA for ALL users, enforced by a full-screen security gate so NO application content, navigation, or data renders until password policy, current Terms acceptance, and MFA are satisfied. (2) 15+ character password policy enforced at four layers. (3) PostgreSQL Row-Level Security for per-row, role-based data isolation. (4) Append-only audit_log and security_events (failed logins, MFA events, idle timeouts, password changes, sensitive-record changes) that cannot be modified/deleted by app roles. (5) 30-minute idle auto-logout for shared clinical workstations. (6) Strict CSP and security headers (X-Frame-Options: DENY, nosniff, HTTPS-only). (7) PHI prohibited by design (zero-tolerance Terms). (8) US-only hosting (AWS us-east-2). (9) Self-service password reset; refresh-token rotation; 1-hour token expiry."),
]

wb = Workbook()
ws = wb.active
ws.title = "Yale Security Review"

# Styles
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
section_fill = PatternFill("solid", fgColor="D9E1F2")
section_font = Font(bold=True, size=11, color="1F4E78")
meta_font = Font(bold=True, size=10)
wrap_top = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title
ws.merge_cells("A1:F1")
ws["A1"] = "Yale New Haven Health System — Data Remote-Hosted Security Design Review"
ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 24

ws.merge_cells("A2:F2")
ws["A2"] = "Vendor: M. Auerbach, MD / B. Michaels (B.L. Michaels, LLC), Yale School of Medicine Pediatrics  |  Application: PECC Support Tool  |  Completed: 06/24/2026  |  'CONFIRM (business)' = vendor must verify org fact before submitting"
ws["A2"].font = Font(italic=True, size=9, color="595959")
ws.row_dimensions[2].height = 16

headers = ["SEC #", "Question", "Yes", "No", "N/A", "Comments"]
hr = 3
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center if c >= 3 and c <= 5 else Alignment(horizontal="left", vertical="center")
    cell.border = border

r = hr + 1
for item in rows:
    sec, question, answer, comment = item
    if sec == SECTION:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        cell = ws.cell(row=r, column=1, value=question)
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = border
        r += 1
        continue

    is_meta = (answer == "" and comment == "" and not str(sec).startswith("SEC"))
    ws.cell(row=r, column=1, value=sec)
    ws.cell(row=r, column=2, value=question)
    yes_mark = "X" if answer == "Yes" else ("Unique" if answer == "Unique" else "")
    no_mark = "X" if answer == "No" else ""
    na_mark = "X" if answer == "N/A" else ""
    # Special-case non Yes/No/NA answers (e.g., "Unique") -> put text in comments lead instead
    if answer == "Unique":
        yes_mark = ""
        comment = "Unique accounts. " + comment
    ws.cell(row=r, column=3, value=yes_mark)
    ws.cell(row=r, column=4, value=no_mark)
    ws.cell(row=r, column=5, value=na_mark)
    ws.cell(row=r, column=6, value=comment)

    ws.cell(row=r, column=1).font = meta_font
    ws.cell(row=r, column=2).alignment = wrap_top
    ws.cell(row=r, column=6).alignment = wrap_top
    for c in (3, 4, 5):
        ws.cell(row=r, column=c).alignment = center
        ws.cell(row=r, column=c).font = Font(bold=True, color="C00000")
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = border
    ws.cell(row=r, column=1).alignment = wrap_top
    r += 1

# Column widths
widths = {"A": 9, "B": 58, "C": 6, "D": 6, "E": 6, "F": 80}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Freeze header
ws.freeze_panes = "A4"

# Auto-ish row heights via wrap (Excel will expand; set a reasonable min)
for row in range(hr + 1, r):
    ws.row_dimensions[row].height = None

wb.save(OUT)
print("WROTE:", OUT)
print("ROWS:", r - hr - 1)
